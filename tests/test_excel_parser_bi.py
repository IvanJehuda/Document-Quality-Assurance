import io

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from excel_parser_bi import MONTH_ABBREVS, _qualify_duplicate_labels, parse_bi_table


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_bi_workbook_bytes(sheet_name="I.1", title="Uang Beredar (M2)", unit="(Miliar Rp)"):
    """Two-year, three-month-each BI table: title/unit/blank/year-row/month-row/2 data rows.

    2023 block occupies cols D-F (Jan-Mar), 2024 block occupies cols G-I (Jan-Mar),
    matching the real TABEL1_1.xls layout (col0=row#, col1=indent, col2=label, col3+=values).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([title])
    ws.append([unit])
    ws.append([])
    ws.append([None, None, None, 2023, None, None, 2024])
    ws.append([None, None, None, "Jan", "Feb", "Mar", "Jan", "Feb", "Mar"])
    ws.append([1, None, "Total", 100.0, 110.0, 120.0, 200.0, 210.0, 220.0])
    ws.append([2, None, "M1", 50.0, 55.0, 60.0, 90.0, 95.0, 100.0])
    return _save(wb)


def test_parse_bi_table_reads_title_unit_and_row_labels():
    result = parse_bi_table(_build_bi_workbook_bytes(), "I.1")

    assert result.title == "Uang Beredar (M2)"
    assert result.unit == "Miliar Rp"
    assert result.row_labels == ["Total", "M1"]


def test_lookup_returns_exact_value_for_each_year_month():
    result = parse_bi_table(_build_bi_workbook_bytes(), "I.1")

    assert result.lookup("Total", 2023, "Jan") == 100.0
    assert result.lookup("Total", 2024, "Mar") == 220.0
    assert result.lookup("M1", 2023, "Feb") == 55.0


def test_lookup_returns_none_for_missing_label_or_period():
    result = parse_bi_table(_build_bi_workbook_bytes(), "I.1")

    assert result.lookup("Nonexistent", 2023, "Jan") is None
    assert result.lookup("Total", 2023, "Dec") is None


def test_lookup_fuzzy_matches_via_substring():
    result = parse_bi_table(_build_bi_workbook_bytes(), "I.1")

    matched_label, value = result.lookup_fuzzy("Tota", 2023, "Jan")

    assert matched_label == "Total"
    assert value == 100.0


def test_lookup_fuzzy_falls_back_to_total_row_when_query_matches_table_subject():
    result = parse_bi_table(_build_bi_workbook_bytes(), "I.1")

    # "Uang Beredar" doesn't match any row label directly, but it does share
    # significant words with the table title "Uang Beredar (M2)".
    matched_label, value = result.lookup_fuzzy("Uang Beredar", 2023, "Jan")

    assert matched_label == "Total"
    assert value == 100.0


def test_lookup_fuzzy_returns_none_when_nothing_matches():
    result = parse_bi_table(_build_bi_workbook_bytes(), "I.1")

    matched_label, value = result.lookup_fuzzy("Completely Unrelated Metric", 2023, "Jan")

    assert matched_label is None
    assert value is None


def test_lookup_fuzzy_prefers_fuller_label_over_short_generic_row_contained_in_query():
    # Regression for BI I.1: the components section has 'Simpanan Berjangka (Rupiah dan Valas)'
    # while the determinants section has a nested row labelled just 'Simpanan' (a negative
    # liability item). 'simpanan' is contained in the query 'simpanan berjangka', and with both
    # containment directions ranked shortest-first in one pool, the generic 'Simpanan' row won
    # and produced a false Refuted. Query-contained-in-label must outrank label-contained-in-query.
    wb = Workbook()
    ws = wb.active
    ws.title = "I.1"
    ws.append(["Uang Beredar dan faktor-faktor yang mempengaruhinya"])
    ws.append(["(Miliar Rp)"])
    ws.append([])
    ws.append([None, None, None, 2026])
    ws.append([None, None, None, "Jan"])
    ws.append([1, None, "Simpanan Berjangka (Rupiah dan Valas)", 3158896.8825])
    ws.append([2, None, "Simpanan", -560955.7467])
    result = parse_bi_table(_save(wb), "I.1")

    matched_label, value = result.lookup_fuzzy("simpanan berjangka", 2026, "Jan")

    assert matched_label == "Simpanan Berjangka (Rupiah dan Valas)"
    assert value == 3158896.8825


def test_lookup_fuzzy_label_in_query_direction_picks_most_specific_label():
    # When only the label-in-query direction matches (verbose LLM query embedding a label),
    # the LONGEST (most specific) label must win, not the shortest.
    wb = Workbook()
    ws = wb.active
    ws.title = "I.1"
    ws.append(["Uang Beredar"])
    ws.append(["(Miliar Rp)"])
    ws.append([])
    ws.append([None, None, None, 2026])
    ws.append([None, None, None, "Jan"])
    ws.append([1, None, "Giro", 1.0])
    ws.append([2, None, "Giro Valas", 760086.4])
    result = parse_bi_table(_save(wb), "I.1")

    matched_label, value = result.lookup_fuzzy("pertumbuhan giro valas pada Januari", 2026, "Jan")

    assert matched_label == "Giro Valas"
    assert value == 760086.4


def test_lookup_fuzzy_does_not_bind_distinct_metric_sharing_a_prefix():
    # A distinct metric that merely shares a long prefix with a real row must NOT be matched:
    # neither name contains the other. (Regression for the old startswith(query[:12]) rule that
    # mis-bound e.g. "Uang Beredar Digital" to "Uang Beredar Luas(M2)" and reported it Refuted.)
    wb = Workbook()
    ws = wb.active
    ws.title = "I.1"
    ws.append(["Uang Beredar dan faktor-faktor yang mempengaruhinya"])
    ws.append(["(Miliar Rp)"])
    ws.append([])
    ws.append([None, None, None, 2026])
    ws.append([None, None, None, "Jan"])
    ws.append([1, None, "Uang Beredar Luas(M2)", 10116181.856])
    result = parse_bi_table(_save(wb), "I.1")

    matched_label, value = result.lookup_fuzzy("Uang Beredar Digital", 2026, "Jan")

    assert matched_label is None
    assert value is None


def test_available_periods_returns_all_periods_for_matched_label():
    result = parse_bi_table(_build_bi_workbook_bytes(), "I.1")

    periods = result.available_periods("Total")

    assert sorted(periods) == sorted(
        [(2023, "Jan"), (2023, "Feb"), (2023, "Mar"), (2024, "Jan"), (2024, "Feb"), (2024, "Mar")]
    )


def test_available_periods_returns_empty_list_when_label_unresolvable():
    result = parse_bi_table(_build_bi_workbook_bytes(), "I.1")

    assert result.available_periods("Nonexistent Metric") == []


def test_year_anchor_at_december_column_still_maps_all_months_correctly():
    # BI tables sometimes anchor the year label at the December cell instead of January -
    # the column-period mapper must back-propagate to find January regardless.
    wb = Workbook()
    ws = wb.active
    ws.title = "I.1"
    ws.append(["Title"])
    ws.append(["(Miliar Rp)"])
    ws.append([])
    year_row = [None, None, None] + [None] * 11 + [2023]
    month_row = [None, None, None] + MONTH_ABBREVS
    values = [float(100 + i) for i in range(12)]
    ws.append(year_row)
    ws.append(month_row)
    ws.append([1, None, "Total"] + values)

    result = parse_bi_table(_save(wb), "I.1")

    assert result.lookup("Total", 2023, "Jan") == 100.0
    assert result.lookup("Total", 2023, "Dec") == 111.0


# ---------------------------------------------------------------------------
# Duplicate-label disambiguation (parent-qualified via the indent hierarchy)
# ---------------------------------------------------------------------------

def test_qualify_duplicate_labels_prefixes_duplicates_with_nearest_parent():
    rows = [
        ("Uang Kuasi", 1, {3: 1.0}),
        ("Simpanan Berjangka", 2, {3: 2.0}),
        ("Rupiah", 3, {3: 3.0}),
        ("Valuta Asing", 3, {3: 4.0}),
        ("Tabungan Lainnya", 2, {3: 5.0}),
        ("Rupiah", 3, {3: 6.0}),
        ("Giro Valas", 2, {3: 7.0}),
    ]

    result = _qualify_duplicate_labels(rows)

    labels = [label for label, _ in result]
    # Duplicated 'Rupiah' gets a parent prefix on EVERY occurrence; unique labels untouched.
    assert labels == [
        "Uang Kuasi",
        "Simpanan Berjangka",
        "Simpanan Berjangka > Rupiah",
        "Valuta Asing",
        "Tabungan Lainnya",
        "Tabungan Lainnya > Rupiah",
        "Giro Valas",
    ]
    # Values stay attached to their own rows.
    assert dict(result)["Simpanan Berjangka > Rupiah"] == {3: 3.0}
    assert dict(result)["Tabungan Lainnya > Rupiah"] == {3: 6.0}


def test_qualify_duplicate_labels_without_indents_degrades_to_bare_names():
    rows = [("Rupiah", 0, {3: 1.0}), ("Rupiah", 0, {3: 2.0})]

    result = _qualify_duplicate_labels(rows)

    assert [label for label, _ in result] == ["Rupiah", "Rupiah"]


def _build_hierarchical_workbook_bytes():
    """Two sections that both contain a 'Rupiah' sub-row with DIFFERENT values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "I.1"
    ws.append(["Uang Beredar dan faktor-faktor yang mempengaruhinya"])
    ws.append(["(Miliar Rp)"])
    ws.append([])
    ws.append([None, None, None, 2026])
    ws.append([None, None, None, "Jan"])
    data = [
        ("Simpanan Berjangka (Rupiah dan Valas)", 2, 3158896.0),
        ("Rupiah", 3, 2748250.0),
        ("Tabungan Lainnya (Rupiah dan Valas)", 2, 2928330.0),
        ("Rupiah", 3, 2696896.0),
        ("Simpanan", 3, -560955.0),
    ]
    for i, (label, indent, value) in enumerate(data, 1):
        ws.append([i, None, label, value])
        ws.cell(row=5 + i, column=3).alignment = Alignment(indent=indent)
    return _save(wb)


def test_parse_bi_table_keeps_each_sections_duplicate_subrow_values():
    result = parse_bi_table(_build_hierarchical_workbook_bytes(), "I.1")

    # Both Rupiah sub-rows survive with their own values (no first-occurrence shadowing).
    assert result.lookup("Simpanan Berjangka (Rupiah dan Valas) > Rupiah", 2026, "Jan") == 2748250.0
    assert result.lookup("Tabungan Lainnya (Rupiah dan Valas) > Rupiah", 2026, "Jan") == 2696896.0


def test_lookup_fuzzy_leaf_tier_uses_parent_words_to_disambiguate():
    result = parse_bi_table(_build_hierarchical_workbook_bytes(), "I.1")

    matched, value = result.lookup_fuzzy("tabungan lainnya rupiah", 2026, "Jan")

    assert matched == "Tabungan Lainnya (Rupiah dan Valas) > Rupiah"
    assert value == 2696896.0


def test_lookup_fuzzy_leaf_tier_beats_generic_label_contained_in_query():
    # 'simpanan' (generic negative row) is contained in the query, but the corroborated
    # leaf match (parent 'Simpanan Berjangka' + leaf 'Rupiah') must win.
    result = parse_bi_table(_build_hierarchical_workbook_bytes(), "I.1")

    matched, value = result.lookup_fuzzy("simpanan berjangka rupiah", 2026, "Jan")

    assert matched == "Simpanan Berjangka (Rupiah dan Valas) > Rupiah"
    assert value == 2748250.0


def test_parse_bi_table_raises_when_sheet_not_found():
    data = _build_bi_workbook_bytes(sheet_name="I.1")

    with pytest.raises(ValueError, match="not found"):
        parse_bi_table(data, "WrongSheet")


def test_parse_bi_table_raises_when_no_year_header_present():
    wb = Workbook()
    ws = wb.active
    ws.title = "I.1"
    ws.append(["Title"])
    ws.append(["(Miliar Rp)"])
    for _ in range(8):
        ws.append(["no year header here"])

    with pytest.raises(ValueError, match="auto-detect year/month header rows"):
        parse_bi_table(_save(wb), "I.1")


def test_parse_bi_table_raises_on_unrecognized_file_format():
    with pytest.raises(ValueError, match="Unrecognized file format"):
        parse_bi_table(b"not an excel file at all", "I.1")
