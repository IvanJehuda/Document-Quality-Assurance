"""Parser for Bank Indonesia (BI) wide statistical tables (e.g. TABEL1_1.xls, sheet I.1).

Table structure (consistent across monthly updates — only new rightmost columns are added):
  Row 0  : Title
  Row 1  : Unit annotation, e.g. "(Miliar Rp)"
  Row 2+ : One or more empty rows (count varies between table series)
  Row N  : Year headers (numeric float, sparse) — auto-detected by scanning for cells
           containing values in the range 1990-2100.
  Row N+1: Month abbreviations (Jan/Feb/.../Dec or Apr* for provisional).
  Row N+2+: Data rows
             col 0 = row index (numeric), col 2 = KETERANGAN label,
             col 3+ = numeric values aligned with (year, month) columns.
             The row hierarchy (which sub-item belongs to which section) is encoded in the
             label CELL's formatting indent level, not in any column's text — col 1 is empty.

Supports .xls (xlrd) and .xlsx (openpyxl). Format detected by magic bytes.
The year/month row positions are auto-detected so tables with different header offsets
(e.g. TABEL I.1 uses rows 3/4 while TABEL 5.9 uses rows 4/5) are handled automatically.

Duplicate labels: BI tables reuse the same sub-item label under different parents
('Pinjaman yang Diberikan 2)' appears once per sector section; 'Rupiah'/'Valuta Asing' under
both Simpanan Berjangka and Tabungan Lainnya). Such duplicates are disambiguated by prefixing
their nearest parent (from the indent hierarchy): 'Tagihan Kepada Sektor Swasta > Pinjaman
yang Diberikan 2)'. Without this, the (label, year, month) data dict silently kept only the
FIRST section's values — observed in production returning loans-to-financial-institutions
figures for a claim about total bank credit.
"""

import io
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

MONTH_ABBREVS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_MONTH_INDEX: Dict[str, int] = {m: i for i, m in enumerate(MONTH_ABBREVS)}

# Minimum column index where data starts (skip row-num, indent, label cols).
_DATA_START_COL = 3

# Separator between a duplicated label and its qualifying parent ("Parent > Child").
QUAL_SEP = " > "


def _sig_words(text: str) -> set:
    """Lowercased words of 3+ chars — the same notion of 'significant' used for title matching."""
    return {w.lower() for w in re.findall(r"\w+", text) if len(w) > 2}


def _normalize_month(raw) -> Optional[str]:
    """Strip trailing * / spaces and return canonical 3-letter abbreviation, or None."""
    if not raw:
        return None
    cleaned = str(raw).strip().rstrip("*").strip()
    return cleaned if cleaned in _MONTH_INDEX else None


def _parse_unit(raw) -> str:
    return str(raw or "").strip().strip("()")


@dataclass
class BITableData:
    """Parsed BI statistical table ready for direct (year, month) lookup."""
    title: str
    unit: str
    row_labels: List[str]

    # Internal: (row_label, year, month_abbrev) -> float  — first occurrence wins for dupes.
    _data: Dict[Tuple[str, int, str], float] = field(default_factory=dict, repr=False)

    def lookup(self, row_label: str, year: int, month: str) -> Optional[float]:
        return self._data.get((row_label, year, month))

    # Words too generic to indicate semantic overlap between a query and the table title.
    _TITLE_STOP_WORDS: frozenset = frozenset({
        "dan", "di", "ke", "dari", "untuk", "yang", "pada", "atau", "dalam",
        "dengan", "oleh", "atas", "total", "posisi", "indonesia", "bank",
        "the", "of", "a", "an", "and", "in", "for",
    })

    def _query_matches_table_subject(self, query: str) -> bool:
        """Return True when significant words in query overlap with the table title.

        Used to decide whether a generic 'Total' row is the right aggregate for
        a query that names the table's overall subject (e.g. query='Cadangan Devisa'
        against a table titled 'Cadangan Devisa Indonesia').
        """
        def sig_words(text: str) -> set:
            return {
                w.lower() for w in re.findall(r'\w+', text)
                if len(w) > 2 and w.lower() not in self._TITLE_STOP_WORDS
            }
        return bool(sig_words(query) & sig_words(self.title))

    def available_periods(self, query: str) -> List[Tuple[int, str]]:
        """Return all (year, month) pairs that have data for the label closest matching query.

        Uses the same tier-1/2/3 label matching as lookup_fuzzy so that diagnostics
        are consistent with what the comparison step would attempt.
        """
        matched = self._resolve_label(query)
        if matched is None:
            return []
        return sorted({(y, m) for (l, y, m) in self._data if l == matched})

    def _match_tiers(self, query: str):
        """Yield candidate row labels for query, best tier first.

        Tier order (each tier yields (label, sort_key) candidates; the caller takes the first
        tier that produces a usable match, picking min(sort_key)):

          1. Case-insensitive exact equality.
          2. Query contained in label — the label is the fuller official name of what the query
             names ("simpanan berjangka" → "Simpanan Berjangka (Rupiah dan Valas)"). SHORTEST
             label wins (closest to the query).
          3. Leaf of a parent-qualified label ('Simpanan Berjangka ... > Rupiah' → 'rupiah')
             contained in the query. Ranked by how many significant words of the FULL
             qualified label appear in the query (desc), so "tabungan lainnya rupiah" picks
             Tabungan Lainnya's Rupiah sub-row over Simpanan Berjangka's, then by leaf
             length. Runs BEFORE the bare label-in-query tier: a leaf match corroborated by
             parent words ("simpanan berjangka valuta asing" → SB's Valuta Asing) must beat
             a short generic label that merely appears in the query ('Simpanan').
          4. Label contained in query — a verbose query embeds an exact label name. LONGEST
             label wins (most specific). This direction is kept LAST and specificity-ranked
             because it is the dangerous one: a short generic row like 'Simpanan' (a nested
             sub-item of a different section) is contained in "simpanan berjangka" and, when
             ranked shortest-first in the same pool as tier 2, shadowed the correct row —
             observed on BI I.1, producing a false Refuted against the negative 'Simpanan'
             liability row.

        Containment (not shared-prefix) in the full-label tiers keeps distinct metrics that
        merely start alike apart: "Uang Beredar Digital" binds to nothing ("Uang Beredar
        Luas(M2)" neither contains it nor is contained by it).
        """
        q_lower = query.lower().strip()
        q_words = _sig_words(q_lower)
        tier_exact = [(label, 0) for label in self.row_labels if label.lower().strip() == q_lower]
        tier_q_in_l = [(label, len(label)) for label in self.row_labels if q_lower in label.lower()]
        tier_l_in_q = [(label, -len(label)) for label in self.row_labels if label.lower().strip() in q_lower]
        tier_leaf = []
        for label in self.row_labels:
            if QUAL_SEP not in label:
                continue
            leaf = label.rsplit(QUAL_SEP, 1)[1].lower().strip()
            if leaf and leaf in q_lower:
                overlap = len(_sig_words(label) & q_words)
                tier_leaf.append((label, (-overlap, -len(leaf))))
        return [tier_exact, tier_q_in_l, tier_leaf, tier_l_in_q]

    def _resolve_label(self, query: str) -> Optional[str]:
        """Return the best matching row label for query, or None if nothing matches."""
        if query in self.row_labels:
            return query
        for tier in self._match_tiers(query):
            if tier:
                return min(tier, key=lambda t: t[1])[0]
        if self.title and self._query_matches_table_subject(query):
            for label in self.row_labels:
                if label.strip().lower() == "total":
                    return label
        return None

    def lookup_fuzzy(
        self, query: str, year: int, month: str
    ) -> Tuple[Optional[str], Optional[float]]:
        """Return (matched_label, value) using exact → tiered containment → Total fallback.

        See _match_tiers for the tier order and why the two containment directions must not
        share one pool. Within a tier, only labels that actually have data for (year, month)
        are considered, so a better-named but data-less row never blocks a usable one.
        The final fallback handles tables where the overall metric (e.g. 'Cadangan Devisa')
        is not a row label but IS the table's subject (from self.title), and the aggregate
        is stored in a row simply called 'Total'.
        """
        # Exact
        v = self._data.get((query, year, month))
        if v is not None:
            return query, v
        for tier in self._match_tiers(query):
            with_data = [
                (label, key) for label, key in tier
                if self._data.get((label, year, month)) is not None
            ]
            if with_data:
                best = min(with_data, key=lambda t: t[1])[0]
                return best, self._data[(best, year, month)]
        # Title-aware Total fallback: query describes this table's subject → return 'Total' row
        if self.title and self._query_matches_table_subject(query):
            for label in self.row_labels:
                if label.strip().lower() == "total":
                    v = self._data.get((label, year, month))
                    if v is not None:
                        return label, v
        return None, None


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _find_header_rows(raw_rows: List[List]) -> Tuple[int, int]:
    """Scan the first 10 rows and return (year_row_idx, month_row_idx).

    The year row is the first row that contains at least one numeric cell
    whose value falls in 1990-2100. The month row is assumed to be the
    immediately following row.
    """
    for i, row in enumerate(raw_rows[:10]):
        if any(isinstance(v, (int, float)) and 1990 <= float(v) <= 2100 for v in row):
            return i, i + 1
    raise ValueError(
        "Could not auto-detect year/month header rows in the first 10 rows. "
        "Expected at least one cell with a numeric year value (1990–2100)."
    )


def _build_col_period_map(
    year_row: List, month_row: List
) -> Dict[int, Tuple[int, str]]:
    """Return {col_index: (year, month_abbrev)} from the raw year/month header rows.

    Year anchors in row 3 can appear at the January OR December cell of a year (BI changed
    convention between historical and recent data).  We back-propagate from the anchor to
    find each year's January column, then assign all 12 months forward from there.
    """
    # --- collect month-labelled columns ----------------------------------------
    month_cols: Dict[int, str] = {}
    for c, val in enumerate(month_row):
        m = _normalize_month(val)
        if m and c >= _DATA_START_COL:
            month_cols[c] = m

    if not month_cols:
        raise ValueError("No month headers found in row 4 of the sheet.")

    sorted_data_cols = sorted(month_cols)

    # --- collect year anchors --------------------------------------------------
    year_anchors: Dict[int, int] = {}
    for c, val in enumerate(year_row):
        if isinstance(val, (int, float)) and 1990 <= float(val) <= 2100:
            year_anchors[c] = int(float(val))

    if not year_anchors:
        raise ValueError("No year headers found in row 3 of the sheet.")

    # --- determine Jan column for each year ------------------------------------
    year_jan_cols: Dict[int, int] = {}
    for anchor_col, year in year_anchors.items():
        month_at_anchor = month_cols.get(anchor_col)
        if month_at_anchor is None:
            continue
        month_idx = _MONTH_INDEX[month_at_anchor]   # 0=Jan … 11=Dec
        if anchor_col not in sorted_data_cols:
            continue
        pos_in_sorted = sorted_data_cols.index(anchor_col)
        jan_pos = pos_in_sorted - month_idx
        if jan_pos >= 0 and year not in year_jan_cols:
            # First (leftmost) occurrence wins — prevents the English mirror section of
            # bilingual BI tables from overwriting the primary Indonesian column mapping.
            year_jan_cols[year] = sorted_data_cols[jan_pos]

    if not year_jan_cols:
        raise ValueError("Could not map any year anchor to a January column.")

    # --- assign (year, month) to every data column -----------------------------
    sorted_years = sorted(year_jan_cols.items(), key=lambda kv: kv[1])   # by col

    col_period_map: Dict[int, Tuple[int, str]] = {}
    for i, (year, jan_col) in enumerate(sorted_years):
        if i + 1 < len(sorted_years):
            next_jan_col = sorted_years[i + 1][1]
            year_cols = [c for c in sorted_data_cols if jan_col <= c < next_jan_col]
        else:
            year_cols = [c for c in sorted_data_cols if c >= jan_col]

        for c in year_cols:
            col_period_map[c] = (year, month_cols[c])

    return col_period_map


# ---------------------------------------------------------------------------
# Duplicate-label disambiguation via the indent hierarchy
# ---------------------------------------------------------------------------

def _qualify_duplicate_labels(
    rows: List[Tuple[str, int, Dict[int, float]]]
) -> List[Tuple[str, Dict[int, float]]]:
    """Disambiguate duplicated labels with their nearest parent from the indent hierarchy.

    rows: (label, indent_level, values) in sheet order. A row's parent is the closest
    preceding row with a smaller indent. Labels that appear more than once in the sheet get
    a 'Parent > Label' qualified name (ALL occurrences, since each is equally ambiguous);
    unique labels are left untouched so common metric names stay stable for the LLM prompt
    and existing lookups. If indents are absent (all zero — e.g. synthetic fixtures or a
    reader that couldn't recover formatting), duplicates have no parent and keep their bare
    name, which degrades to the old first-occurrence-wins behaviour.
    """
    counts: Dict[str, int] = {}
    for label, _, _ in rows:
        counts[label] = counts.get(label, 0) + 1

    stack: List[Tuple[int, str]] = []  # (indent, bare label) of open ancestors
    out: List[Tuple[str, Dict[int, float]]] = []
    for label, indent, vals in rows:
        while stack and stack[-1][0] >= indent:
            stack.pop()
        parent = stack[-1][1] if stack else None
        final = f"{parent}{QUAL_SEP}{label}" if counts[label] > 1 and parent else label
        stack.append((indent, label))
        out.append((final, vals))
    return out


# ---------------------------------------------------------------------------
# .xls reader (xlrd)
# ---------------------------------------------------------------------------

def _read_xls(data: bytes, sheet_name: str):
    import xlrd  # optional dependency

    # formatting_info is needed to read each label cell's indent level (the only place BI
    # tables encode the row hierarchy). Fall back to no-formatting parsing when unavailable.
    try:
        wb = xlrd.open_workbook(file_contents=data, formatting_info=True)
        has_formatting = True
    except Exception:
        wb = xlrd.open_workbook(file_contents=data)
        has_formatting = False
    try:
        ws = wb.sheet_by_name(sheet_name)
    except xlrd.XLRDError:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheet_names()}")

    title = str(ws.cell_value(0, 0)).strip()
    unit = _parse_unit(ws.cell_value(1, 0))

    scan_rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(min(10, ws.nrows))]
    year_row_idx, month_row_idx = _find_header_rows(scan_rows)
    year_row = scan_rows[year_row_idx]
    month_row = scan_rows[month_row_idx]
    col_period_map = _build_col_period_map(year_row, month_row)

    def indent_of(r: int) -> int:
        if not has_formatting:
            return 0
        try:
            return int(wb.xf_list[ws.cell_xf_index(r, 2)].alignment.indent_level or 0)
        except Exception:
            return 0

    raw_rows: List[Tuple[str, int, Dict[int, float]]] = []
    for r in range(month_row_idx + 1, ws.nrows):
        label = str(ws.cell_value(r, 2)).strip()
        if not label:
            continue
        if re.match(r'^\d+\)', label) or len(label) > 120:
            break
        vals: Dict[int, float] = {}
        for c in col_period_map:
            raw = ws.cell_value(r, c)
            if isinstance(raw, (int, float)):
                vals[c] = float(raw)
        raw_rows.append((label, indent_of(r), vals))

    return title, unit, col_period_map, _qualify_duplicate_labels(raw_rows)


# ---------------------------------------------------------------------------
# .xlsx reader (openpyxl)
# ---------------------------------------------------------------------------

def _read_xlsx(data: bytes, sheet_name: str):
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    # Cells (not values_only) so the label cell's alignment indent — the row-hierarchy
    # signal — is available alongside the value. ReadOnlyCell exposes .alignment.
    cell_rows = list(wb[sheet_name].iter_rows())
    all_rows = [[c.value for c in row] for row in cell_rows]

    title = str(all_rows[0][0] or "").strip() if all_rows else ""
    unit = _parse_unit(all_rows[1][0] if len(all_rows) > 1 else "")

    scan_rows = [list(all_rows[r]) for r in range(min(10, len(all_rows)))]
    year_row_idx, month_row_idx = _find_header_rows(scan_rows)
    year_row = scan_rows[year_row_idx]
    month_row = scan_rows[month_row_idx]
    col_period_map = _build_col_period_map(year_row, month_row)

    def indent_of(r_idx: int) -> int:
        try:
            cell = cell_rows[r_idx][2]
            return int(getattr(cell.alignment, "indent", 0) or 0)
        except Exception:
            return 0

    raw_rows: List[Tuple[str, int, Dict[int, float]]] = []
    for r_idx in range(month_row_idx + 1, len(all_rows)):
        row = all_rows[r_idx]
        if len(row) < 3:
            continue
        label = str(row[2] or "").strip()
        if not label:
            continue
        if re.match(r'^\d+\)', label) or len(label) > 120:
            break
        vals: Dict[int, float] = {}
        for c in col_period_map:
            if c < len(row) and isinstance(row[c], (int, float)):
                vals[c] = float(row[c])
        raw_rows.append((label, indent_of(r_idx), vals))

    return title, unit, col_period_map, _qualify_duplicate_labels(raw_rows)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def list_sheet_names(data: bytes) -> List[str]:
    """Return the worksheet names in an .xls/.xlsx workbook (format by magic bytes).

    Used to populate the sheet picker in the UI so users choose from the actual
    sheets instead of typing a name. Reads only workbook metadata (fast).
    """
    if data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        import xlrd  # optional dependency, only needed for legacy .xls
        return list(xlrd.open_workbook(file_contents=data, on_demand=True).sheet_names())
    elif data[:4] == b'PK\x03\x04':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    else:
        raise ValueError("Unrecognized file format: expected .xls or .xlsx bytes.")


def parse_bi_table(data: bytes, sheet_name: str) -> BITableData:
    """Parse a BI statistical table from .xls or .xlsx bytes.

    Args:
        data: Raw file bytes.
        sheet_name: Sheet to parse, e.g. "I.1".

    Returns:
        BITableData supporting .lookup(row_label, year, month) and .lookup_fuzzy().
    """
    # Detect format by magic bytes
    if data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        title, unit, col_period_map, row_data = _read_xls(data, sheet_name)
    elif data[:4] == b'PK\x03\x04':
        title, unit, col_period_map, row_data = _read_xlsx(data, sheet_name)
    else:
        raise ValueError("Unrecognized file format: expected .xls or .xlsx bytes.")

    result = BITableData(title=title, unit=unit, row_labels=[])

    for label, vals in row_data:
        result.row_labels.append(label)
        for col, (year, month) in col_period_map.items():
            if col in vals:
                key = (label, year, month)
                # Duplicated labels arrive parent-qualified from the readers, so collisions
                # here are rare (same label twice under the same-named parent, or a reader
                # that couldn't recover indents). First occurrence wins for those leftovers.
                if key not in result._data:
                    result._data[key] = vals[col]

    return result
